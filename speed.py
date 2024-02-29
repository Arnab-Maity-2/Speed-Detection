import cv2
import os
import pandas as pd
from ultralytics import YOLO
from tracker import *
import time
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from math import dist
model = YOLO('yolov8s.pt')
countc = 0

def RGB(event, x, y, flags, param):
    if event == cv2.EVENT_MOUSEMOVE:
        colorsBGR = [x, y]
        # print(colorsBGR)


def save_frame(image, x3, y3, x4, y4):
    # Calculate new coordinates for the cropped region
    padding = 100
    x3_cropped = max(0, x3 - padding)
    y3_cropped = max(0, y3 - padding)
    x4_cropped = min(image.shape[1], x4 + padding+50)
    y4_cropped = min(image.shape[0], y4 + padding)

    # Crop the image
    cropped_image = image[y3_cropped:y4_cropped, x3_cropped:x4_cropped]

    return cropped_image

def excel_data(filename_ex="excel_data.xlsx", sheetname="vehicle_info", data=None):
    try:
        # Load the workbook
        # workbook = openpyxl.load_workbook(filename_ex)

        # # Get the worksheet by name
        # worksheet = workbook[sheetname]

        # Find the next available row
        next_row = worksheet.max_row + 1
        # worksheet.row_dimensions[next_row].height = 200

        # Insert serial number in the first column
        worksheet.cell(row=next_row, column=1, value=next_row - 1)
        worksheet.row_dimensions[next_row].height = 100


        # Insert data in subsequent columns
        for col, value in enumerate(data, start=2):
            worksheet.cell(row=next_row, column=col, value=value)
            img = Image(image_name)
            img.width = 125  # Adjust the width of the image
            img.height = 100  # Adjust the height of the image
            # Place the image in the 6th column of the specified row
            img.anchor = f'{get_column_letter(6)}{next_row}'
            worksheet.add_image(img)



        # Save the changes
        print("New row added successfully.")

    except FileNotFoundError:
        print(f"File '{filename_ex}' not found.")


with open('stored.txt', 'w') as file:
    pass


def append_text(filename, text, speed):
    with open(filename, 'a') as file:
        if speed > 35:
            file.write(text + ' (overspeeding)' + '\n')
            vio = "overspeeding"
        elif speed < 20:
            file.write(text + ' (underspeeding)' + '\n')
            vio = "underspeeding"
        else:
            file.write(text + '\n')
            vio = False
    return vio

filename_ex = "excel_data.xlsx"
sheetname = "vehicle_info"
Headers = ['#', 'Direction', 'Sl_no.', 'Speed', 'Violation', 'Photos']

if os.path.isfile(filename_ex):
    os.remove(filename_ex)
else:
    pass


workbook = openpyxl.Workbook()
worksheet = workbook.create_sheet(title=sheetname, index=0)
workbook.save(filename_ex)
workbook = openpyxl.load_workbook(filename_ex)
worksheet = workbook[sheetname]
for col, value in enumerate(Headers, start=1):
    worksheet.cell(row=1, column=col, value=value)
workbook.save(filename_ex)


cv2.namedWindow('RGB')
cv2.setMouseCallback('RGB', RGB)

cap = cv2.VideoCapture('veh3.mp4')


my_file = open("coco.txt", "r")
data = my_file.read()
class_list = data.split("\n")
# print(class_list)

output_dir = "violator_vehicles"
for item in os.listdir(output_dir):
    item_path = os.path.join(output_dir, item)
    if os.path.isfile(item_path):
        # Delete files
        os.remove(item_path)
os.makedirs(output_dir, exist_ok=True)

count = 0

tracker = Tracker()

cy1 = 322
cy2 = 368

offset = 6

vh_down = {}
counter = []


vh_up = {}
counter1 = []

while True:
    ret, frame = cap.read()
    if not ret:
        break
    count += 1
    if count % 3 != 0:
        continue
    frame = cv2.resize(frame, (1020, 500))

    results = model.predict(frame)
 #   print(results)
    a = results[0].boxes.data
    px = pd.DataFrame(a).astype("float")
#    print(px)
    list = []

    for index, row in px.iterrows():
        #        print(row)

        x1 = int(row[0])
        y1 = int(row[1])
        x2 = int(row[2])
        y2 = int(row[3])
        d = int(row[5])
        c = class_list[d]
        if 'car' in c or 'motorcycle' in c:
            list.append([x1, y1, x2, y2])
    bbox_id = tracker.update(list)
    for bbox in bbox_id:
        x3, y3, x4, y4, id = bbox
        cx = int(x3+x4)//2
        cy = int(y3+y4)//2

        cv2.rectangle(frame, (x3, y3), (x4, y4), (0, 0, 255), 2)

        if cy1 < (cy+offset) and cy1 > (cy-offset):
            vh_down[id] = time.time()
            cv2.circle(frame, (cx, cy), 4, (0, 0, 255), -1)
        if id in vh_down:

            if cy2 < (cy+offset) and cy2 > (cy-offset):
                cv2.circle(frame, (cx, cy), 4, (255, 0, 0), -1)
                elapsed_time = time.time() - vh_down[id]
                if counter.count(id) == 0:
                    counter.append(id)
                    distance = 10  # meters
                    a_speed_ms = distance / elapsed_time
                    a_speed_kh = a_speed_ms * 3.6
                    # cv2.circle(frame, (cx, cy), 4, (0, 0, 255), -1)
                    cv2.putText(frame, str(id), (x3, y3),
                                cv2.FONT_HERSHEY_COMPLEX, 0.6, (255, 255, 255), 1)
                    cv2.putText(frame, str(int(a_speed_kh))+'Km/h', (x4, y4),
                                cv2.FONT_HERSHEY_COMPLEX, 0.8, (0, 255, 255), 2)
                    text = f"goingdown : {len(counter)} Speed : {a_speed_kh} km/hr"
                    vio = append_text('stored.txt', text, a_speed_kh)
                    if vio != False:
                        image_name = os.path.join(
                            output_dir, f"snapshot_{countc}.jpg")
                        cv2.imwrite(image_name, save_frame(
                            frame, x3, y3, x4, y4))
                        # image = Image(image_name)?
                        # print(f"Snapshot saved: {image_name}")
                        countc += 1

                        data = ['Going Down', len(counter), str(
                            round(a_speed_kh, 2))+'km/hr', vio]
                        excel_data(data=data)

        ##### going UP#####
        if cy2 < (cy+offset) and cy2 > (cy-offset):
            vh_up[id] = time.time()
            # cv2.circle(frame, (cx, cy), 4, (0, 0, 255), -1)
        if id in vh_up:

            if cy1 < (cy+offset) and cy1 > (cy-offset):
                # cv2.circle(frame, (cx, cy), 4, (0, 255, 255), -1)
                elapsed1_time = time.time() - vh_up[id]

                if counter1.count(id) == 0:
                    counter1.append(id)
                    distance1 = 10  # meters
                    a_speed_ms1 = distance1 / elapsed1_time
                    a_speed_kh1 = a_speed_ms1 * 3.6
                    cv2.circle(frame, (cx, cy), 4, (255, 0, 0), -1)
                    cv2.putText(frame, str(id), (x3, y3),
                                cv2.FONT_HERSHEY_COMPLEX, 0.6, (255, 255, 255), 1)
                    cv2.putText(frame, str(int(a_speed_kh1))+'Km/h', (x4, y4),
                                cv2.FONT_HERSHEY_COMPLEX, 0.8, (0, 255, 255), 2)
                    text = f"goingup   : {len(counter1)} Speed : {a_speed_kh1} km/hr"
                    vio = append_text('stored.txt', text, a_speed_kh1)
                    if vio != False:
                        image_name = os.path.join(
                            output_dir, f"snapshot_{countc}.jpg")
                        cv2.imwrite(image_name, save_frame(
                            frame, x3, y3, x4, y4))
                        # print(f"Snapshot saved: {image_name}")
                        # image = Image(image_name)
                        countc += 1

                        data = ['Going Up', len(counter1), str(
                            round(a_speed_kh1, 2)) + "km/hr", vio]
                        excel_data(data=data)

    cv2.line(frame, (00, cy1), (1020, cy1), (255, 255, 255), 1)

    cv2.putText(frame, ('L1'), (15, cy1-2),
                cv2.FONT_HERSHEY_COMPLEX, 0.8, (0, 255, 255), 2)

    cv2.line(frame, (00, cy2), (1020, cy2), (255, 255, 255), 1)

    cv2.putText(frame, ('L2'), (15, cy2-2),
                cv2.FONT_HERSHEY_COMPLEX, 0.8, (0, 255, 255), 2)
    d = (len(counter))
    u = (len(counter1))
    cv2.putText(frame, ('goingdown:-')+str(d), (60, 90),
                cv2.FONT_HERSHEY_COMPLEX, 0.8, (0, 255, 255), 2)

    cv2.putText(frame, ('goingup:-')+str(u), (60, 130),
                cv2.FONT_HERSHEY_COMPLEX, 0.8, (0, 255, 255), 2)
    cv2.imshow("RGB", frame)
    if cv2.waitKey(0) & 0xFF == 27:
        break


for col in worksheet.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
            # Add some padding and scaling factor
    adjusted_width = (max_length + 2) * 1.1
    worksheet.column_dimensions[column].width = adjusted_width
worksheet.column_dimensions['f'].width = 20
for row in worksheet.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

workbook.save(filename_ex)
workbook.close()
cap.release()
cv2.destroyAllWindows()
